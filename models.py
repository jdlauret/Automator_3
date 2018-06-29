# %% Imports
import datetime as dt
import json
import os
import re
import subprocess as sp

import cx_Oracle
import numpy as np
import pandas as pd
import pygsheets
from time import sleep
from dateutil.parser import parse
from openpyxl import Workbook, load_workbook
from pydrive.auth import GoogleAuth
from pydrive.drive import GoogleDrive
from utilities.recurrences import recur_test

# %% Module Directories
top_dir = os.path.dirname(__file__)
credentials_dir = os.path.join(top_dir, 'credentials')
script_dir = os.path.join(top_dir, 'script_storage')
file_storage_dir = os.path.join(top_dir, 'file_storage')


# %% General Use Functions
def check_for_file(file_path):
    """
    Check if a file exists in target folder
    :param file_path: The File Path including the File Name
    :return: Boolean
    """
    for file_name in os.listdir(os.path.dirname(file_path)):
        if file_name in os.path.basename(file_path):
            return True
    return False


def colnum_string(n):
    """
    Converts a number to a letter value
    If the number is greater than 26 it will increment to AA, AB, etc
    :param n: The number to convert to letter
    :return: String of Letter Value
    """
    div = n
    string = ""
    while div > 0:
        mod = (div - 1) % 26
        string = chr(65 + mod) + string
        div = int((div - mod) / 26)
    return string


def range_builder(start_row, start_col, end_row=None, end_col=None):
    """
    Builds A1 Notation using the number values
    Start Row and Column are required
    End Row and Column are optional
    start_row 1 + start_col 1 = A1
    start_row 1 + start_col 1 + end_col 2 = A1:B
    start_row 1 + start_col 1 + end_row 2 + end_col 2 = A1:B2
    :param start_row: Starting Row Number
    :param start_col: Starting Column Number
    :param end_row: Ending Row Number (Cannot be used alone like end_col)
    :param end_col: Ending Column Number
    :return: A1 Notation String
    """
    # Uses colnum_string to return a letter value
    start_col_letter = colnum_string(start_col)

    # If both End row and col are not None return Start Cell and Cell "A1:B2"
    if end_row is not None and end_col is not None:
        end_col_letter = colnum_string(end_col)
        range_name = str(start_col_letter) + str(start_row) + ':' + str(end_col_letter) + str(end_row)
    # If only End Col is not None return Start Cell and End Col "A1:B"
    elif end_row is None and end_col is not None:
        end_col_letter = colnum_string(end_col)
        range_name = str(start_col_letter) + str(start_row) + ':' + str(end_col_letter)
    # If both end row and col are None return Single Cell Notation "A1"
    else:
        range_name = str(start_col_letter) + str(start_row)
    return range_name


# %% Data Warehouse
class DataWarehouse:
    """
    DataWarehouse class is used for interacting with the Vivint Solar Data Warehouse
    """

    # Universal Character set to remove from strings and replace with a regular space
    bad_characters = [chr(9), chr(10), chr(13)]

    def __init__(self, user, connection_type='prod', encoding='utf-8'):
        """
        Initialize DataWarehouse Class
        :param user: User key to access credentials
        :param connection_type: Connect to the Production or Dev
        :param encoding: Select the Type of encoding for data coming out
                         of and into the data warehouse
        """
        # Set Submitted Parameters
        self.user = user
        self.connection_type = connection_type
        self.encoding = encoding

        # Setup all basic variables
        self.results = None
        self.column_data = None
        self.column_names = []
        self.data_for_upload = None
        self.results_with_header = None
        self.batch_errors = []
        self.failed_to_format = []

        # Used to ignore the id column when sending information to the data warehouse
        self.ignore_id_col = True

        self.db = None
        self.cursor = None

    def open_connection(self):
        """
        Open credentials file and create a new connection to the Database
        """
        connection_info = json.loads(os.environ.get('DATA_WAREHOUSE'))

        # Values retrieved from credentials file
        user = connection_info['credentials'][self.user]
        host = connection_info['connections'][self.connection_type].get('host')
        port = connection_info['connections'][self.connection_type].get('port')
        sid = connection_info['connections'][self.connection_type].get('sid')

        # Setup DNS_TNS connection info
        dns_tns = cx_Oracle.makedsn(host, port, sid)

        try:
            # Open Connection
            self.db = cx_Oracle.connect(user.get('username'), user.get('password'), dns_tns,
                                        encoding=self.encoding, nencoding=self.encoding)
        except cx_Oracle.DatabaseError as e:
            # TODO create Exception Handling
            # Log error as appropriate
            raise

        # If the database connection succeeded create the cursor for use.
        self.cursor = self.db.cursor()

    def close_connection(self):
        """
        Close Database Connection
        """
        try:
            self.cursor.close()
            self.db.close()
        except cx_Oracle.DatabaseError:
            raise

    def execute(self, sql, bindvars=None, commit=True):
        """
        Executes a SQL statement
        Example bindvars {"myField": aValue, "anotherOne": anotherValue}
        :param sql: SQL statement to Execute
        :param bindvars: Dictionary of variables passed to execute.
        :param commit: Commit to Database
        """

        self.open_connection()

        # Format SQL Statement for better results
        query = self.format_query(sql)
        try:
            # Execute Statement
            self.cursor.execute(query, bindvars)

            if commit:
                # Commit to Database
                self.db.commit()
        except cx_Oracle.DatabaseError as e:
            # TODO create Exception Handling
            # Log error as appropriate
            raise

        self.close_connection()

    def set_results_with_header(self):
        """
        Adds column_names to the top of results
        """
        self.results_with_header = self.column_names + self.results

    def update_task(self, task_id, column_name, new_value, table_name='JDLAURET.T_AUTO_TASKS'):
        query = """UPDATE {table_name}
                SET {column_name} = {new_value}
                WHERE ID = {task_id}""".format(table_name=table_name,
                                               column_name=column_name,
                                               new_value=new_value,
                                               task_id=task_id)
        self.open_connection()
        try:
            self.cursor.execute(query)
        except:
            # TODO create Exception Handling (probably as it's own function)
            pass
        self.close_connection()

    def update_last_run(self, task_id, run_time):
        table_name = 'JDLAURET.T_AUTO_TASKS'
        column_name = 'LAST_RUN'
        last_run = run_time.strftime('%m-%d-%y %H:%M:%S')
        query = """UPDATE {table_name} t
                        SET t.{column_name} = TO_DATE({new_value}, 'MM-DD-YY HH24:MI:SS')
                        WHERE t.ID = {task_id}""".format(table_name=table_name,
                                                       column_name=column_name,
                                                       new_value=last_run,
                                                       task_id=task_id)
        self.open_connection()
        try:
            self.cursor.execute(query)
        except:
            # TODO create Exception Handling (probably as it's own function)
            pass
        self.close_connection()

    def get_columns_data(self, table_name, ignore_id_col=True):
        """
        Sets column data for the table name submitted
        :param table_name: schema.table_name to access
        :param ignore_id_col: set self.ignore_id_col value
        """
        self.ignore_id_col = ignore_id_col
        # Select Statement
        query = "SELECT * FROM {0} WHERE 1=0".format(table_name)

        self.open_connection()
        # Execute Query
        self.cursor.execute(query)
        # Retrieve Column Data and convert to list of tuples
        self.column_data = list(self.cursor.description)
        # If self.ignore_id_col is True create column names minus the 'ID' Column
        # Otherwise set all column names
        if self.ignore_id_col:
            self.column_names = [x[0] for x in self.column_data if x[0].lower() != 'id']
        else:
            self.column_names = [x[0] for x in self.column_data]

    def format_query_results(self):
        """
        Converts all LOB types to strings for easier use
        """
        for i, row in enumerate(self.results):
            for j, cell in enumerate(row):
                if isinstance(cell, cx_Oracle.LOB):
                    self.results[i][j] = str(cell)

    def query_results(self, sql, bindvars=None):
        """
        Sets self.results to the results of the SQL query
        :param sql: SQL Query to be executed
        :param bindvars: Dictionary of variables passed to execute.
        """
        self.open_connection()
        # Formats query to remove help execution
        query = self.format_query(sql)
        try:
            if bindvars is not None:
                # Prepare the Query for execution
                self.cursor.prepare(query)
                # Execute Query with bindvars
                self.cursor.execute(query, bindvars)
            else:
                # Execute Query without bindvars
                self.cursor.execute(query)
            # Set column data, column names, and results
            self.column_data = list(self.cursor.description)
            self.column_names = [x[0] for x in self.column_data]
            self.results = [list(x) for x in self.cursor.fetchall()]
            self.format_query_results()
            self.set_results_with_header()

        except cx_Oracle.DatabaseError as e:
            # TODO Create Exception Handling
            raise

        finally:
            self.close_connection()

    def format_query(self, sql):
        """
        Format a SQL Query to make running it with cx_Oracle a little easier
        :param sql: SQL Query to format
        :return: String of formatted SQL Query
        """
        return re.sub(' +', ' ', sql).replace(';', '')

    def get_table_data(self, table_name):
        """
        Get all data from a table
        Set self.results to list of lists containing entire table data
        :param table_name: schema.table_name
        """
        query = "SELECT * FROM {table_name}".format(table_name=table_name)
        self.query_results(query)

    def remove_characters(self, string):
        """
        Remove characters from a string
        :param string: String to be evaluated
        :return: String with removed characters
        """
        new_string = string
        for word in self.bad_characters:
            new_string.replace(word, ' ')
        return new_string

    def format_data(self):
        """
        Loop through data that is going into the Data Warehouse and format the data
        This step helps with mis-matched data types in columns
        """

        # loop through every row and column of data
        for i, row in enumerate(self.data_for_upload):

            for j, item in enumerate(row):
                # get current column type (STRING, NUMBER, DATE, etc.)
                column_type = self.column_data[j][1].__name__
                # Set all Nones to Blank string, cx_Oracle will turn this into a null value
                if item is None:
                    self.data_for_upload[i][j] = ''

                # Evaluate all non-blank cells
                if item != '':
                    # If item not a datetime type attempt to parse and create datetime object
                    if column_type == 'DATETIME' \
                            or column_type == 'TIMESTAMP':
                        if not isinstance(item, dt.datetime):
                            try:
                                self.data_for_upload[i][j] = parse(item).strftime("%Y-%m-%d %H:%M:%S")
                            except (TypeError, ValueError) as e:
                                self.failed_to_format.append([self.data_for_upload[i], e])
                        else:
                            try:
                                self.data_for_upload[i][j] = item.strftime("%Y-%m-%d %H:%M:%S")
                            except (TypeError, ValueError) as e:
                                self.failed_to_format.append([self.data_for_upload[i], e])
                    # If item not String change to string
                    if column_type == 'STRING' or \
                            column_type == 'LOB':
                        if not isinstance(item, str):
                            self.data_for_upload[i][j] = str(item)
                        self.data_for_upload[i][j] = self.remove_characters(self.data_for_upload[i][j])
                    # if item not integer convert to integer
                    if column_type == 'NUMBER':
                        self.data_for_upload[i][j] = str(item)

    def insert_data_to_table(self, table_name, data, column_map=False, ignore_id_col=True, header_included=False):
        """
        Insert data into a table
        :param table_name: schema.table to insert data into
        :param data: A list of lists to insert
        :param column_map: A list of the Columns in the order they are associated with the data
        :param ignore_id_col: If something is being inserted into the ID col set to False
        :param header_included: If the first line in the data is the header set to True
        """
        self.ignore_id_col = ignore_id_col

        self.data_for_upload = data
        self.get_columns_data(table_name)

        # Setup column map for upload
        if column_map:
            self.column_data = [x for x in self.column_data if x[0] in column_map]
            self.column_names = column_map

        self.format_data()

        # Remove header from data set
        if header_included:
            del data[0]

        # Set data size variables
        data_row_count = len(data)
        data_column_count = max(len(x) for x in data)
        print('Data Column Count:', data_column_count)
        print('Column Map Count:', len(self.column_names))

        # Check if there is more data columns than columns in the table
        if data_column_count != len(self.column_names):
            raise ValueError('Number of Columns submitted does not match Table or Column Map')

        if data_row_count > 0:
            # Rows of data found
            bind_variables = []
            # setup required variables for Insert Statement
            for i, column_name in enumerate(self.column_names):
                new_bind_name = ':' + column_name
                class_name = self.column_data[i][1].__name__
                if class_name == 'TIMESTAMP':
                    new_bind_name = 'TO_TIMESTAMP(:{0}, \'{1}\')'.format(column_name, 'yyyy-mm-dd hh24:mi:ss')
                    bind_variables.append(new_bind_name)
                elif class_name == 'DATETIME':
                    new_bind_name = 'TO_DATE(:{0}, \'{1}\')'.format(column_name, 'yyyy-mm-dd hh24:mi:ss')
                    bind_variables.append(new_bind_name)
                else:
                    bind_variables.append(new_bind_name)
            print(', '.join(bind_variables))
            # Format Insert Statement
            query = """INSERT INTO {table_name} ({column_names}) VALUES ({values})
                    """.format(table_name=table_name,
                               column_names=', '.join(self.column_names),
                               values=', '.join(bind_variables))
            self.open_connection()
            # Prepare Query and setup bind names
            self.cursor.prepare(query)
            self.cursor.bindnames()

            try:
                # Execute Insert Statement
                self.cursor.executemany(None, self.data_for_upload,
                                        batcherrors=True)
                # Store Batch Errors
                self.batch_errors = self.cursor.getbatcherrors()
                number_of_errors = len(self.batch_errors)
                if number_of_errors > 0:
                    print('{0} batch errors encountered'.format(number_of_errors))
                    print('Error Preview')
                    for j in range(number_of_errors):
                        print(self.batch_errors[j].message)
                        if j > 2:
                            break
                # Commit To Data base
                self.db.commit()

            except:
                # TODO setup Exception handling
                raise

            self.close_connection()


# %% CSV Generation
class CsvGenerator:
    """
    Generate CSV in File Storage directory
    """

    def __init__(self, data, params):
        """
        Setup for csv
        The params should be handled by the Task class automatically
        However here is the basic Dictionary Needs
        { "header": [List of items for the name of each column],
          "file_name": "name_of_file.csv", - The File name for the csv
          "file_path": "path\\to\\file, - The File Path to store the file in
          "dynamic_name": "%y-%m-%d", - This is a date format in python datetime formatting
          "after_before": "after" - If the dynamic name should show up before or after the file name
        }
        :param data: A list of lists to put in a csv
        :param params: A dictionary containing all needed information
        """

        self.data = data
        self.params = params
        self.header = self.params.get('header')
        self.file_path = file_storage_dir
        self.file_name = self.params.get('file_name')
        self.dynamic_name = params.get('dynamic_name')

        self.successful_run = False

        # Remove the header from the data set
        # if it is included in the data set
        if self.header is None:
            self.header = data[0]
            del self.data[0]

        # Add the .csv to the end of the file name
        # if it not included
        if '.csv' not in self.file_name:
            self.file_name = self.file_name + '.csv'

        # If dynamic_name is included set dynamic name
        if self.dynamic_name is not None:
            self.create_dynamic_name()

    def create_dynamic_name(self):
        """
        Create new Dynamic File Name
        """
        before_after = self.params.get('after_before')
        # Use before or after in after_before to setup new file name
        if before_after.lower() == 'after':
            self.file_name = self.file_name.replace('.csv', '').strip() \
                             + ' ' \
                             + str(dt.datetime.strftime(dt.datetime.today(), self.dynamic_name)).strip() \
                             + '.csv'
        else:
            self.file_name = str(dt.datetime.strftime(dt.datetime.today(), self.dynamic_name)).strip() \
                             + ' ' \
                             + self.file_name

    def create_csv(self):
        """
        Write data to csv using pandas
        """
        try:
            # Convert List of Lists to DataFrame and write it to a CSV
            pd.DataFrame(self.data, columns=self.header).to_csv(os.path.join(self.file_path, self.file_name))
            self.successful_run = True
        except:
            # TODO create Exception Handling
            raise


# %% Python Scripts
class PythonScript:
    """
    PythonScript used to run additional python scripts
    """

    def __init__(self, params):
        """
        :param params: The task_data dictionary from Task()
        """
        self.file_name = params.get('file_name')

        scripts = os.listdir(script_dir)

        if self.file_name in scripts:
            self.file_path = script_dir
        else:
            self.file_path = params.get('data_source_id')

        self.script_path = os.path.join(self.file_path, self.file_name)
        self.successful_run = False

    def run_script(self):
        """
        Call python script and read Return Codes
        """
        try:
            # Generate command console python command
            command = 'python' + ' "' + self.script_path + '"'

            child = sp.Popen(command)
            stream_data = child.communicate()[0]

            # rc is the return code of the script
            rc = child.returncode

            # Check return code
            if rc == 0:
                self.successful_run = True

        except Exception as e:
            # TODO create Exception Handling
            raise e


# %% Excel Generator
class ExcelGenerator:
    """
    Generate Excel file with Data submitted
    """

    def __init__(self, data, params):
        """
        Setup for Excel
        The params should be handled by the Task class automatically
        However here is the basic Dictionary Needs
        { "header": [List of items for the name of each column],
          "file_name": "name_of_file.csv", - The File name for the csv
          "file_path": "path\\to\\file, - The File Path to store the file in
          "dynamic_name": "%y-%m-%d", - This is a date format in python datetime formatting
          "after_before": "after", - If the dynamic name should show up before or after the file name
          "wb_sheet_name": "Sheet1", - The sheet to put the data in
          "range_params": { "wb_start_row": 2, - The starting row for the data to go on
                            "wb_start_column": 1, - The starting column for the data
                            "wb_end_row": None or int, - This is optional and can be used to limit the range of the data
                            "wb_end_column": None or int, - Also optional refer to end_row
                          } ,
          "data_storage_id": "", - If there is a master file this is the ID of the item in Google Drive
          ""
        }
        :param data: A list of lists to put in a csv
        :param params: A dictionary containing all needed information
        """
        self.params = params

        self.data = data
        self.data_len = 0
        self.data_wid = 0

        self.wb = None
        self.ws = None

        # Extract values from params dict
        self.header = params.get('header')

        self.file_name = params.get('file_name')
        self.dynamic_name = params.get('dynamic_name')

        self.sheet_name = params.get('wb_sheet_name')

        self.range_params = params.get('range_params')
        self.start_row = self.range_params.get('wb_start_row')
        self.start_col = self.range_params.get('wb_start_column')
        self.end_row = self.range_params.get('wb_end_row')
        self.end_col = self.range_params.get('wb_end_column')

        self.download_id = self.params.get('data_storage_id')

        self.file_path = os.path.join(file_storage_dir, self.file_name)

        # If no header is included in params dict
        # Assumes the header is part of the Data Set
        if self.header is None:
            self.header = self.data[0]
            del self.data[0]

        if self.dynamic_name is not None:
            self.create_dynamic_name()

        # Setup range name for inserting data
        self.range_name = None
        self.create_range()

    def create_dynamic_name(self):
        """
        Create new Dynamic File Name
        """
        before_after = self.params.get('after_before')
        # Use before or after in after_before to setup new file name
        if before_after.lower() == 'after':
            self.file_name = self.file_name.replace('.xlsx', '').strip() \
                             + ' ' \
                             + str(dt.datetime.strftime(dt.datetime.today(), self.dynamic_name)).strip() \
                             + '.xlsx'
        else:
            self.file_name = str(dt.datetime.strftime(dt.datetime.today(), self.dynamic_name)).strip() \
                             + ' ' \
                             + self.file_name

    def create_range(self):
        """
        Creates an A1 Notation of range
        Uses the start row and column and the data width and length
        to determine full range name.
        """
        # Set Data length and width
        self.data_len = len(self.data)
        self.data_wid = max(len(x) for x in self.data)

        # Evaluate end row and col for numbers and adjusts
        # their values based on start row and col values if needed
        if self.end_row is not None:
            self.end_row = int(self.end_row) + (self.start_row - 1)
        else:
            self.end_row = self.data_len + (self.start_row - 1)

        if self.end_col is not None:
            self.end_col = int(self.end_col) + (self.start_col - 1)
        else:
            self.end_col = self.data_wid + (self.start_col - 1)

        self.range_name = range_builder(self.start_row, self.start_col, end_row=self.end_row, end_col=self.end_col)

    def create_excel(self):
        """
        Creates an excel file in the desired location
        """
        if len(self.data) > 0:
            try:
                # If download ID exists, check if there is an existing file present
                # If no existing file is present download the file
                if self.download_id is not None:
                    # Download Existing File
                    if not check_for_file(self.file_path):
                        gdrive_params = {
                            'file_id': self.download_id,
                            'file_name': self.file_name,
                            'file_path': self.file_path
                        }
                        GDrive(params=gdrive_params).download_file()
            except:
                # TODO create Exception Handling
                # If no existing file pass to workbook creation to build new workbook
                pass

            try:
                # Open/Create Workbook
                self.create_workbook()

                # Mark data_storage_complete as True
                data_storage_complete = True
            except Exception as e:
                # TODO emailer function or bug reporting goes here
                # Email if Workbook creation fails
                # send_email(task_name, 'Data Storage: Create Excel Workbook', e)
                pass

    def create_workbook(self):
        """
        Creates workbook and writes data to the workbook
        """
        print()
        print("Opening Workbook")

        try:
            if '.xlsm' in self.file_name or '.xltm' in self.file_name:
                self.wb = load_workbook(self.file_path, keep_vba=True)
            else:
                if '.xlsx' not in self.file_name:
                    self.file_name = self.file_name + '.xlsx'
                self.wb = load_workbook(os.path.join(self.file_path, self.file_name))
        except Exception as e:
            print('Failed to open {0}'.format(self.file_name))
            self.wb = Workbook()

        print("Workbook Opened")

        print('Creating Sheet \'{0}\''.format(self.sheet_name))
        sheet_names = self.wb.sheetnames
        if self.sheet_name in sheet_names:
            self.ws = self.wb.get_sheet_by_name(self.sheet_name)
        else:
            self.ws = self.wb.create_sheet(title=self.sheet_name)

        if len(self.data) < self.ws.max_row:
            self.clear_sheet()
            self.wb.save(self.file_path)

        self.write_to_sheet()
        print('Saving Workbook')

        self.wb.save(os.path.join(self.file_path, self.file_name))

    def clear_sheet(self):
        for row in self.ws[self.range_name]:
            for cell in row:
                cell.value = None

    def write_to_sheet(self):
        """
        Writes a list to the worksheets named
        """
        print("Writing Data to {0}".format(self.ws))

        for i, row in enumerate(self.ws[range]):
            if isinstance(self.data[i], tuple):
                self.data[i] = list(self.data[i])
        print()
        print('Data written to {0}'.format(self.ws))


# %% Google Sheets
class GSheets:
    def __init__(self, sheet_id):
        self.sheet_id = self.key_extractor(sheet_id)
        self.credentials = self.get_credentials()
        self.gc = pygsheets.authorize(self.credentials, outh_creds_store='credentials')
        self.spreadsheet = self.gc.open_by_key(self.sheet_id)

    def key_extractor(self, url):
        drive_key_len = 28
        spreadsheet_key_len = 44
        if url is not None:
            if len(url) > drive_key_len \
                    or len(url) > spreadsheet_key_len:

                if 'file/d/' in url:
                    file_start = url.index('file/d/') + len('file/d/')
                    new_url = url[file_start:]
                    return new_url[:new_url.index('/')]

                if 'spreadsheets/d/' in url:
                    file_start = url.index('spreadsheets/d/') + len('spreadsheets/d/')
                    new_url = url[file_start:]
                    return new_url[:new_url.index('/')]

                if 'folders/' in url:
                    file_start = url.index('folders/') + len('folders/')
                    return url[file_start:]

                if 'id=' in url:
                    file_start = url.index('id=') + len('id=')
                    return url[file_start:]

        return url

    def get_credentials(self):
        print(os.getcwd())
        return 'credentials/client_secret.json'

    def get_row_count(self, sheet_name):
        wks = self.spreadsheet.worksheet('title', sheet_name)
        return wks.rows

    def get_column_count(self, sheet_name):
        wks = self.spreadsheet.worksheet('title', sheet_name)
        return wks.cols

    def add_rows(self, sheet_name, number_of_rows):
        wks = self.spreadsheet.worksheet('title', sheet_name)
        wks.add_rows(number_of_rows)

    def add_cols(self, sheet_name, number_of_columns):
        wks = self.spreadsheet.worksheet('title', sheet_name)
        wks.add_cols(number_of_columns)

    def get_sheet_data(self, sheet_name):
        wks = self.spreadsheet.worksheet('title', sheet_name)
        return wks.get_all_values()

    def clear_sheet_data(self, sheet_name, start, end):
        wks = self.spreadsheet.worksheet('title', sheet_name)
        wks.clear(start=start, end=end)

    def update_sheet(self, sheet_name, start, data_for_upload):
        wks = self.spreadsheet.worksheet('title', sheet_name)

        if not isinstance(data_for_upload, pd.DataFrame):
            df = pd.DataFrame(data_for_upload)
        else:
            df = data_for_upload

        df = df.replace(np.nan, '', regex=True)
        wks.set_dataframe(df, start, copy_head=False)

    def update_named_range(self, sheet_name, range_name, value):
        wks = self.spreadsheet.worksheet('title', sheet_name).update_cell(range_name, value)

    def get_named_range(self, sheet_name, range_name):
        wks = self.spreadsheet.worksheet('title', sheet_name)
        range_data = wks.get_named_range(range_name)
        return range_data


# %% Google Drive
class GDrive:

    def __init__(self, params=None):
        self.gAuth = GoogleAuth()

        self.params = params
        self.file_name = self.params.get('file_name')
        self.file_path = self.params.get('file_path')
        self.folder_id = self.params.get('folder_id')
        self.file_id = self.params.get('file_id')
        self.cred_file = os.path.join(credentials_dir, 'mycreds.txt')

        self.drive = self.g_auth()

        if self.folder_id is not None:
            self.folder_id = self.folder_id
        else:
            self.folder_id = 'root'

        self.file_list = self.list_files()

        if self.file_id is None:
            for item in self.file_list:
                if item['title'] == self.file_name:
                    self.file_id = item.get('id')
        if self.file_path is not None and self.file_name is not None:
            self.file = os.path.join(self.file_path, self.file_name)

    def g_auth(self):
        self.gAuth.LoadCredentialsFile(self.cred_file)
        if self.gAuth.credentials is None:
            # Authenticate if they're not there
            self.gAuth.LocalWebserverAuth()
        elif self.gAuth.access_token_expired:
            # Refresh them if expired
            self.gAuth.Refresh()
        else:
            # Initialize the saved creds
            self.gAuth.Authorize()
            # Save the current credentials
            self.gAuth.SaveCredentialsFile(self.cred_file)

        return GoogleDrive(self.gAuth)

    def list_files(self):
        return self.drive.ListFile({'q': "'{folder_id}' in parents and trashed=false"
                                   .format(folder_id=self.folder_id)}).GetList()

    def start_upload(self):
        if self.file_id is None:
            f = self.drive.CreateFile(metadata={
                'title': self.file_name,
                "parents": [
                    {
                        "kind": "drive#fileLink",
                        "id": self.folder_id
                    }
                ]
            })
        else:
            f = self.drive.CreateFile(metadata={
                'id': self.file_id,
                'title': self.file_name,
                "parents": [
                    {
                        "kind": "drive#fileLink",
                        "id": self.folder_id
                    }
                ]
            })

        f.SetContentFile(self.file)
        f.Upload()

    def download_file(self):
        f = self.drive.CreateFile({'id': self.file_id})
        mime_type = f['mimeType']
        f.GetContentFile(os.path.join(self.file_path, self.file_name), mimetype=mime_type)

    def get_file_name(self):
        return str(self.drive.CreateFile({'id': str(self.file_id)})['title'])

    def read_drive_file(self):
        f = self.drive.CreateFile({'id': self.file_id})
        return f.GetContentString()


# %% Task Metrics
class TaskMetrics:
    def __init__(self, params):
        self.id = params.get('id')
        self.run_type = params.get('operational')
        self.input_time = params.get('input_time')
        self.output_time = params.get('ouput_time')
        self.upload_time = params.get('upload_time')
        self.task_completion_time = 0
        self.upload_line = [[self.id, self.task_completion_time, self.input_time,
                             self.output_time, self.upload_time, 0,
                             self.run_type, dt.datetime.now()
                             ]]
        self.dw = DataWarehouse('admin')
        self.table_name = 'JDLAURET.T_AUTO_METRICS'
        self.set_task_completion_time()

    def set_task_completion_time(self):
        if isinstance(self.input_time, int):
            self.task_completion_time += self.input_time
        if isinstance(self.output_time, int):
            self.task_completion_time += self.output_time
        if isinstance(self.upload_time, int):
            self.task_completion_time += self.upload_time

    def submit_task_time(self):
        self.dw.insert_data_to_table(self.table_name, self.upload_line, header_included=False, ignore_id_col=False)


# %% Task
class Task:
    def __init__(self, task_line, task_header, run_type='Automated'):
        self.dw = DataWarehouse('admin')
        self.task_line = task_line
        self.task_header = task_header
        self.run_type = run_type
        self.task_data = {}
        self.metrics = None

        self.task_name = None
        self.task_id = None
        self.run_requested = None
        self.data_source = None
        self.data_source_id = None
        self.data_storage = None
        self.data_storage_id = None
        self.storage_id = None

        self.operational = None
        self.file_name = None

        self.input_complete = False
        self.output_complete = False
        self.task_complete = False

        self.input_data_header = None
        self.input_data = None

        self.range_start = 'A2'
        self.range_end = None
        self.range_name = None

        self.metrics = None

        self.failed_attempts = 0

        self.refresh_task_data()

    def refresh_task_data(self):

        query = 'SELECT * FROM JDLAURET.T_AUTO_TASKS T WHERE T.ID = :ID'
        if self.task_id is not None:
            self.dw.query_results(query, bindvars=[self.task_id])
            self.task_line = self.dw.results[0]
            self.task_header = self.dw.column_names

        self.create_task_object()

        self.task_name = self.task_data.get('namex')
        self.task_id = self.task_data.get('id')
        self.run_requested = self.task_data.get('run_requested')
        self.data_source = self.task_data.get('data_source')
        self.data_storage = self.task_data.get('data_storage')
        self.data_source_id = self.task_data.get('data_source_id')
        self.data_storage_id = self.task_data.get('data_storage_id')
        self.storage_type = self.task_data.get('storage_type')
        self.storage_id = self.task_data.get('storage_id')
        self.operational = self.task_data.get('operational')

    def create_task_object(self):
        for i, item in enumerate(self.task_header):
            if self.task_line[i] == '':
                self.task_line[i] = None

            self.task_data[str(item.lower())] = self.task_line[i]
        self.task_data['run_type'] = self.run_type

    def get_input_data(self):
        input_time_start = dt.datetime.now()
        if self.data_source.lower() == 'sql':
            dw = DataWarehouse('admin')
            query = GDrive({'file_id': self.data_source_id}).read_drive_file()
            dw.query_results(query)

            self.input_data = dw.results
            self.input_data_header = dw.column_names
            self.task_data['header'] = self.input_data_header
            self.input_complete = True

            if len(self.input_data) > 0:
                self.set_output_data()

        elif self.data_source.lower() == 'sql command':
            query = GDrive({'file_id': self.data_source_id}).read_drive_file()
            self.dw.execute(query)
            self.task_complete = True

        elif self.data_source.lower() == 'python':
            p_script = PythonScript(self.task_data)
            p_script.run_script()
            self.task_complete = p_script.successful_run

        self.task_data['input_time'] = (dt.datetime.now() - input_time_start).seconds

    def create_range(self):
        data_len = len(self.input_data)
        data_wid = max(len(x) for x in self.input_data)
        start_row = self.task_data['wb_start_row']
        start_col = self.task_data['wb_start_column']
        end_row = self.task_data['wb_end_row']
        end_col = self.task_data['wb_end_column']

        if end_row is not None:
            end_row = int(end_row) + (start_row - 1)
        else:
            end_row = data_len + (start_row - 1)

        if end_col is not None:
            end_col = int(end_col) + (start_col - 1)
        else:
            end_col = data_wid + (start_col - 1)

        self.range_start = range_builder(start_row, start_col)
        self.range_end = range_builder(end_row, end_col)
        self.range_name = range_builder(start_row, start_col, end_row=end_row, end_col=end_col)

    def set_output_data(self):
        output_time_start = dt.datetime.now()
        if self.input_complete:
            if self.data_storage == 'csv':
                try:
                    csv = CsvGenerator(self.input_data, self.task_data)
                    csv.create_csv()
                    self.file_name = csv.file_name
                    self.output_complete = True
                except:
                    # TODO Create Exception
                    pass

            elif self.data_storage == 'excel':
                try:
                    excel = ExcelGenerator(self.input_data, self.task_data)
                    excel.create_excel()
                    self.file_name = excel.file_name
                    self.output_complete = True
                except:
                    # TODO Create Exception
                    pass

            elif self.data_storage == 'google sheets':
                self.create_range()
                sheet_name = self.task_data['wb_sheet_name']
                try:
                    gs = GSheets(self.data_storage_id)
                    gs.clear_sheet_data(sheet_name, self.range_start, self.range_end)
                    gs.update_sheet(sheet_name, self.range_start, self.input_data)
                    self.output_complete = True
                except:
                    # TODO Create Exception
                    pass
            self.task_data['output_time'] = (dt.datetime.now() - output_time_start).seconds

    def upload_files(self):
        upload_time_start = dt.datetime.now()
        if self.storage_type == 'Google Drive':
            for file in os.listdir(file_storage_dir):
                if file == self.file_name:
                    drive = GDrive(params=self.task_data)
                    drive.file_path = file_storage_dir
                    drive.start_upload()
                    self.task_complete = True
        self.task_data['upload_time'] = (dt.datetime.now() - upload_time_start).seconds

    def update_last_run(self):
        if self.run_requested.lower() != 'true':
            task_id = self.task_data.get('id')

            self.dw.update_last_run(task_id, dt.datetime.now())

    def update_run_requested(self):
        self.dw.update_task(self.task_id, 'RUN_REQUESTED', 'FALSE')

    def run_task(self):
        self.refresh_task_data()
        if (recur_test(self.task_data)
            or self.run_requested.lower() == 'true') \
                and self.operational != 'Non-Operational':
            print(self.task_name, 'in progress')
            self.get_input_data()
            if self.input_complete:
                self.set_output_data()
                if self.output_complete:
                    self.upload_files()

            if self.task_complete:
                if self.run_requested.lower() == 'true':
                    self.update_run_requested()
                self.update_last_run()
                print(self.task_name, 'completed')
                self.metrics = TaskMetrics(self.task_data)
                self.metrics.submit_task_time()


# %% FailedRun
class FailedRun:
    def __init__(self, task_id):
        pass


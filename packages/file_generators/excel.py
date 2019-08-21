from . import *
from openpyxl import Workbook, load_workbook


# %% Excel Generator
class ExcelGenerator:
    """
    Generate Excel file with Data submitted
    """

    def __init__(self, data, file_name, sheet_name, range, file_path=None, dynamic_name=None):
        self.data = data
        if self.data:
            self.data_len = len(self.data)
            self.data_wid = max(len(x) for x in self.data)

        self.wb = None
        self.ws = None

        self.file_name = file_name
        self.sheet_name = sheet_name
        self.range_name = range

        self.file_path = file_path

    def create_workbook(self):
        """
        Creates workbook and writes data to the workbook
        """
        try:
            if '.xlsm' in self.file_name or '.xltm' in self.file_name:
                self.wb = load_workbook(self.file_path, keep_vba=True)
            else:
                if '.xlsx' not in self.file_name:
                    self.file_name = self.file_name + '.xlsx'
                self.wb = load_workbook(os.path.join(self.file_path, self.file_name))
        except Exception as e:
            self.wb = Workbook()

        sheet_names = self.wb.sheetnames
        if self.sheet_name in sheet_names:
            self.ws = self.wb[self.sheet_name]
        else:
            self.ws = self.wb.create_sheet(title=self.sheet_name)

        if self.data_len < self.ws.max_row:
            self.clear_sheet()
            self.wb.save(os.path.join(self.file_path, self.file_name))

        if self.data_len:
            self.write_to_sheet()
            self.wb.save(os.path.join(self.file_path, self.file_name))

    def clear_sheet(self):
        for row in self.ws[self.range_name]:
            for cell in row:
                cell.value = None

    def write_to_sheet(self):
        """
        Writes a list to the worksheets named
        """
        for i, row in enumerate(self.ws[self.range_name]):
            if isinstance(self.data[i], tuple):
                self.data[i] = list(self.data[i])
            for j, cell in enumerate(row):
                cell.value = self.data[i][j]
import multiprocessing as mp
import datetime as dt
import sys
import csv
import os
import json
from BI.data_warehouse.connector import Snowflake
from openpyxl import Workbook, load_workbook
sys.path.append(os.path.join(os.path.dirname(__file__), '..'))
from word_list_generator import list_generator
from download_qualtrics_data import DownloadSurveyType
from qualtrics_processor import qualtrics_processor, format_survey_data


def find_data_file(filename):
    if getattr(sys, 'frozen', False):
        # The application is frozen
        data_dir = os.path.dirname(sys.executable)
    else:
        # The application is not frozen
        # Change this bit to match where you store your data files:
        data_dir = os.path.dirname(__file__)

    return os.path.join(data_dir, filename)


def find_main_dir():
    if getattr(sys, 'frozen', False):
        # The application is frozen
        return os.path.dirname(sys.executable)
    else:
        # The application is not frozen
        # Change this bit to match where you store your data files:
        return os.path.dirname(__file__)


def clear_temp_data():
    """
    Deletes all files contained in temp_files
    """
    print('Clearing Temp Data')
    temp_files = [f for f in os.listdir('temp_data')]
    for f in temp_files:
        os.remove(os.path.join('temp_data', f))


def clear_qualtrics_data():
    """
    Deletes all files contained in temp_files
    """
    print('Clearing Temp Data')
    temp_files = [f for f in os.listdir('qualtrics_data')]
    for f in temp_files:
        if f != 'desktop.ini':
            os.remove(os.path.join('qualtrics_data', f))


class CreateQualtricsJson:

        def __init__(self, file_name):
            self.file_name = file_name

        def create_lists_from_excel(self):
            """
            Opens each CSV file in the temp_data directory
            Then stores the list in a json
            :return:
            """
            fn = 'qualtrics_data\\' + self.file_name
            with open(fn, 'r', encoding='utf8') as f:
                new_list = list(list(rec) for rec in csv.reader((
                    line.replace('\0', '') for line in f), delimiter=','))
            new_file_name = self.file_name.replace('csv', 'json')
            fn = 'data\\qualtrics_jsons\\' + new_file_name
            try:
                file = open(fn, 'r')
            except IOError:
                file = open(fn, 'w')
            file.close()
            file = open(fn, 'w')
            new_obj = {str(new_file_name.replace('.json', '')): new_list}
            json.dump(new_obj, file, sort_keys=True, indent=4)
            file.close()


def get_file(file_name, folder_name):
    """
    Creates file string
    :param file_name:
    :param folder_name:
    :return: returns concatenated string
    """
    return folder_name + '/' + file_name


def get_list(key):
    """
    Returns a list from word_and_id json using key
    :param key:
    :return:
    """
    with open(get_file('word_and_id.json', 'data')) as infile:
        word_and_id = json.load(infile)

    word_list = [
        [
            'Survey ID',
            key,
            'NPS Type'
        ]
    ]

    for survey_id in word_and_id.keys():

        for word in word_and_id[survey_id][key]['ENPS']['Word_List']:
            nps_row = [survey_id, word, 'ENPS']
            word_list.append(nps_row)

    return word_list


def print_progress(iteration, total, prefix='', suffix='', decimals=1, bar_length=100):
    """
    Call in a loop to create terminal progress bar
    @params:
        iteration   - Required  : current iteration (Int)
        total       - Required  : total iterations (Int)
        prefix      - Optional  : prefix string (Str)
        suffix      - Optional  : suffix string (Str)
        decimals    - Optional  : positive number of decimals in percent complete (Int)
        bar_length  - Optional  : character length of bar (Int)
    """
    str_format = "{0:." + str(decimals) + "f}"
    percents = str_format.format(100 * (iteration / float(total)))
    filled_length = int(round(bar_length * iteration / float(total)))
    bar = '█' * filled_length + '-' * (bar_length - filled_length)

    sys.stdout.write('\r%s |%s| %s%s %s' % (prefix, bar, percents, '%', suffix)),

    if iteration == total:
        sys.stdout.write('\n')
    sys.stdout.flush()


def write_to_sheet(arr, ws):
    """
    Writes a list to the worksheets named
    :param arr:
    :param ws:
    :return:
    """
    print("Writing Data to {0}".format(ws))
    print_progress(0, len(arr), prefix='Progress', suffix='Complete')
    for r, line in enumerate(arr):
        try:
            ws.append(line)
        except:
            if isinstance(line, tuple):
                arr[r] = list(line)
            ws.append(arr[r])
        print_progress(r, len(arr), prefix='Progress', suffix='Complete')

    print('Data written to {0}'.format(ws))


def create_workbook(arr, wb_name, ws_name, folder_location=None):
    """
    Creates workbook and writes data to the workbook
    :param arr: Desired list or lists to be written to the worksheet
    :param wb_name: Name of the Workbook
    :param ws_name: Name of the Worksheet to write to
    :return:
    """
    print()
    print("Opening Workbook")
    if folder_location is not None:
        os.chdir(folder_location)

    try:
        wb = load_workbook(wb_name)
    except:
        wb = Workbook()

    print("Workbook Opened")

    print('Creating Sheet \'{0}\''.format(ws_name))
    if ws_name in wb.sheetnames:
        sheet_to_remove = wb.get_sheet_by_name(ws_name)
        wb.remove_sheet(sheet_to_remove)
        wb.save(wb_name)

    ws = wb.create_sheet(title=ws_name)
    write_to_sheet(arr, ws)
    print('Saving Workbook')

    wb.save(wb_name)

    if folder_location is not None:
        os.chdir(MAIN_DIR)
    print('Workbook Saved')


def combine_list(list_1, list_2):
    """
    Combine 2 list of lists together removing the header from the second list
    :param list_1:
    :param list_2:
    :return:
    """
    del list_2[0]
    return list_1 + list_2


survey_dict = {
    'exit_2017': {
        'name': 'Exit 2017 (New)',
        'sid': "SV_0x24srHabajaRXn",
        'api_token': os.environ.get('QUALTRICS_HR_API_TOKEN_2'),
        'data_center': 'co1'
    },
    'onboarding_2017': {
        'name': 'Oboarding 2017 (New)',
        'sid': "SV_6o2ccgTXfhSpYAR",
        'api_token': os.environ.get('QUALTRICS_HR_API_TOKEN_2'),
        'data_center': 'co1'
    },
    'ee_mar_hist_2017': {
        'name': 'March 2017 Employee Engagement (Historical - Pulled from Research Suite)',
        'sid': "SV_8ALSvUiqYBqd57T",
        'api_token': os.environ.get('QUALTRICS_HR_API_TOKEN_2'),
        'data_center': 'co1'
    },
    'ee_mar_2018': {
        'name': 'March 2018 Employee Engagement VSLR',
        'sid': "SV_5p7X2Fo5b9xMZgh",
        'api_token': os.environ.get('QUALTRICS_HR_API_TOKEN_2'),
        'data_center': 'co1'
    },
    'turnover_aug_2016': {
        'name': 'Turnover Report for EE August 2016',
        'sid': "SV_0GQMKtGNlaWPseF",
        'api_token': os.environ.get('QUALTRICS_HR_API_TOKEN'),
        'data_center': 'az1'
    },
    'ee_mar_2017': {
        'name': 'March 2017 Employee Engagement',
        'sid': "SV_2uAEveZ25dWtgZD",
        'api_token': os.environ.get('QUALTRICS_HR_API_TOKEN'),
        'data_center': 'az1'
    },
    'eps': {
        'name': 'Employee Pulse Survey',
        'sid': "SV_6s62fOiONKqkNxz",
        'api_token': os.environ.get('QUALTRICS_HR_API_TOKEN'),
        'data_center': 'az1'
    },
    'eps_jun_2016': {
        'name': 'Employee Pulse Survey - June 2016',
        'sid': "SV_1QU03tlkhharrSd",
        'api_token': os.environ.get('QUALTRICS_HR_API_TOKEN'),
        'data_center': 'az1'
    },
    'ee_aug_2016': {
        'name': 'Aug 2016 Employee Engagement - Attrition Copy',
        'sid': "SV_bEKVyX2XnbCgW3z",
        'api_token': os.environ.get('QUALTRICS_HR_API_TOKEN'),
        'data_center': 'az1'
    },
    'feps_q3_2016': {
        'name': 'Final Employee Pulse Survey - Q3 2016 w/Gender',
        'sid': "SV_0e4qt0P0mRtZu4Z",
        'api_token': os.environ.get('QUALTRICS_HR_API_TOKEN'),
        'data_center': 'az1'
    },
    'ee_aug_2016_copy': {
        'name': 'Aug 2016 Employee Engagement - Attrition & Region Copy',
        'sid': "SV_6QXTMMIB7lA20Qt",
        'api_token': os.environ.get('QUALTRICS_HR_API_TOKEN'),
        'data_center': 'az1'
    },
    'rts-om__feedback_loop': {
        'name': 'RTS - Field Feedback',
        'sid': "SV_8eUFJC9DEyzqdQF",
        'api_token': os.environ.get('QUALTRICS_HR_API_TOKEN'),
        'data_center': 'az1'
    }
}


def worker(arg):
    obj, meth_name = arg[:2]
    return getattr(obj, meth_name)(*arg[2:])


if __name__ == '__main__':
    get_new_data = True
    update_tables = True
    words_reset = False

    db = Snowflake()
    db.set_user("JDLAURET")

    NUM_CORE = 3
    MAIN_DIR = find_main_dir()
    TEMP_FOLDER = find_data_file('temp_files')
    DATA_FOLDER = find_data_file('data')

    pool1 = mp.Pool(NUM_CORE)

    # Dates
    start = dt.datetime.now()
    today = dt.datetime.now()
    today = str(today)

    if get_new_data:
        clear_qualtrics_data()
        print('Downloading Qualtrics Data')
        survey_ids = [
            [
                survey_dict[key]['sid'],
                survey_dict[key]['name'],
                survey_dict[key]['data_center'],
                survey_dict[key]['api_token'],
            ] for key in survey_dict.keys()]
        survey_objects = [DownloadSurveyType(line[0], line[1], line[2], line[3]) for line in survey_ids]
        pool1.map(worker, ((obj, 'get_qualtrics_data') for obj in survey_objects))
        # for obj in survey_objects:
        #     obj.get_qualtrics_data()

        files_for_extraction = [filename for filename in os.listdir('qualtrics_data') if filename.endswith('.csv')]
        extraction_objects = [CreateQualtricsJson(file_name) for file_name in files_for_extraction]

        print('Extracting Qualtrics Survey Data')
        pool1.map(worker, ((obj, 'create_lists_from_excel') for obj in extraction_objects))
        # for obj in extraction_objects:
        #     obj.create_lists_from_excel()
        print('Survey Data Processed and Saved')
        pool1.close()
        pool1.join()

    ENPS_TABLE = 'D_POST_INSTALL.T_ENPS_SURVEY_RESPONSE'
    ONBOARDING_TABLE = 'D_POST_INSTALL.T_ENPS_ONBOARDING'
    EXIT_TABLE = 'D_POST_INSTALL.T_ENPS_EXIT'
    NPS_FILE_ID = '0B9Fc6ijLP56VbmFGc0V5MzZOU00'

    print('Reformatting Qualtrics Data')
    data_warehouse_data = qualtrics_processor('data_warehouse')
    data_warehouse_data = format_survey_data(data_warehouse_data)
    onboarding_data = qualtrics_processor('onboarding')
    exit_data = qualtrics_processor('exit')
    print('Qualtrics formatted and in memory')

    # NPS Dash - Projects v2.sql

    word_cloud_table = 'D_POST_INSTALL.T_ENPS_WORD_CLOUD_DATA'
    print('Download Complete')

    reviewed_data_file = 'reviewed_surveys'
    data_dir = os.getcwd() + '\\data'
    workbook_name = 'ENPS Survey Data.xlsx'

    word_cloud_data = get_list(list_generator(data_warehouse_data, data_warehouse_data[0], 1, reset=words_reset))
    # Reset will clear relevant JSON files when a reset of the data is required.
    if update_tables:
        try:
            db.open_connection()
            db.insert_into_table(word_cloud_table, word_cloud_data, overwrite=True, header_included=True)
            db.insert_into_table(ENPS_TABLE, data_warehouse_data, overwrite=True, header_included=True)
            db.insert_into_table(ONBOARDING_TABLE, onboarding_data, overwrite=True, header_included=True)
            db.insert_into_table(EXIT_TABLE, exit_data, overwrite=True, header_included=True)
        finally:
            db.close_connection()

    end = dt.datetime.now()
    duration = end - start
    print('All Tasks Completed in {0} seconds'.format(duration.total_seconds()))

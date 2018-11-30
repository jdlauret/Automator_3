import datetime as dt
import os
import sys
import json
import pandas as pd
import pyexcel
from time import sleep
from openpyxl import load_workbook, Workbook
from BI.data_warehouse import SnowflakeV2, SnowflakeConnectionHandlerV2
from BI.web_crawler import CrawlerBase
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC


def find_main_dir():
    """
    Returns the Directory of the main running file
    """
    if getattr(sys, 'frozen', False):
        # The application is frozen
        return os.path.dirname(sys.executable)

    else:
        # The application is not frozen
        # Change this bit to match where you store your data files:
        return os.path.dirname(os.path.realpath(__file__))


# Directory Variables
MAIN_DIR = find_main_dir()
DOWNLOAD_DIR = os.path.join(MAIN_DIR, 'downloads')

# REPORT_URL points to report page
REPORT_URL = 'https://bu4595898.nicewfm.incontact.com/group/supervisor-webstation/reports-view'

# End point Table name
TABLE_NAME = 'JDLAURET.T_ENPHASE_PORTAL_H'

# MOVE PAYLOAD TO SEPARATE FILE WHEN TESTING IS COMPLETE
payload = {
    'username': {
        'html_name': '_58_login',
        'value': os.environ.get('WEBSTATION_USER')
    },
    'password': {
        'html_name': '_58_password',
        'value': os.environ.get('WEBSTATION_PASS')
    },
}


def save_json(file_name, json_dict):
    """
    Saves an obj to a json
    :param file_name: name of the file to dump the data into
    :param json_dict: the object to dump
    """
    with open(file_name, 'w') as outfile:
        json.dump(json_dict, outfile, indent=4)


def open_json(file_name):
    """
    Opens a json file and returns it as a dictionary
    :param file_name: file_path to the json file
    :return: dictionary of json
    """
    with open(file_name) as infile:
        return json.load(infile)


def extract_number(string_with_number):
    """
    Find all numbers in a string and put them in an array
    :param string_with_number: A string with numbers to be extracted
    :return: First number found in string
    """
    numbers = [int(s) for s in str.split(string_with_number) if s.isdigit()]
    return numbers[0]


def rename_report(num):
    """
    Rename report.xls files to Report - num.xls
    :param num: The number that should be put in the report name
    """
    file_list = os.listdir(DOWNLOAD_DIR)
    file_name = os.path.join(DOWNLOAD_DIR, 'report - ' + str(num) + '.xls')
    for file in file_list:
        file_path = os.path.join(DOWNLOAD_DIR, file)
        if file == 'report.xls':
            os.rename(file_path, file_name)


def clear_downloads():
    """
    Delete all files in the download list
    """
    file_list = os.listdir(DOWNLOAD_DIR)
    for file in file_list:
        if file != 'desktop.ini':
            file_path = os.path.join(DOWNLOAD_DIR, file)
            os.remove(file_path)


class PrimaryCrawler(CrawlerBase):
    """
    Primary Crawler is a web crawler that will login to the Login URL
    """
    num_reports = 3
    crawler_log = {
        'login': False,
        'number_of_downloads': 0
    }

    def __init__(self, driver_type, download_directory=None, headless=False):
        # Primary Items
        CrawlerBase.__init__(self, driver_type, download_directory=download_directory, headless=None)
        # Primary Items
        self.current_line = 0
        self.skip = False

    def run_crawler(self):
        payload = {
            'url': 'https://bu4595898.nicewfm.incontact.com',
            'username': {'input': os.environ.get('WEBSTATION_USER'), 'name': '_58_login',},
            'password': {'input': os.environ.get('WEBSTATION_PASS'), 'name': '_58_password',},
            'submit': {'class': 'aui-button-input',},
        }
        # Login to page
        self.login(**payload)

        # Go to Reports Page
        self.driver.get(REPORT_URL)
        WebDriverWait(self.driver, self.delay) \
            .until(EC.presence_of_element_located((By.ID, '_supvreportsview_WAR_supvwebstationportlet_iframe')))
        # Find iFrame and switch to it
        self.driver.switch_to.frame(self.driver.find_element_by_id('_supvreportsview_WAR_supvwebstationportlet_iframe'))
        try:
            # Wait for the Today table to load and define it
            WebDriverWait(self.driver, self.delay).until(EC.presence_of_element_located((By.ID, 'Today')))
            today_table = self.driver.find_element_by_id('Today')
            new_table = []
            # Get the rows from the table
            rows = today_table.find_elements(By.TAG_NAME, 'tr')

            # Iterate through rows, find links and text and store them
            for row in rows:
                new_row = []
                cols = row.find_elements(By.TAG_NAME, 'td')
                for col in cols:
                    if col.text.lstrip() == 'Adherence' or col.text.lstrip() == 'Agent Schedules':
                        link = col.find_element(By.TAG_NAME, 'a').get_attribute('href')
                        new_row.append(link)
                    new_row.append(col.text.lstrip().rstrip())
                new_table.append(new_row)
            del new_table[0]

            column_names = ['Blank1', 'Links', 'Report_Name', 'Blank2', 'Description', 'Start', 'End', 'Generated',
                            'Size']
            # Convert table to DataFrame Sort by Generation date
            new_table = pd.DataFrame(new_table, columns=column_names)
            new_table = new_table.drop('Blank1', axis=1)
            new_table = new_table.drop('Blank2', axis=1)
            new_table['Generated'] = pd.to_datetime(new_table['Generated'], format="%m/%d/%y %I:%M %p")
            new_table = new_table.sort_values('Generated', ascending=False)
            new_table = new_table.iloc[:self.num_reports]

            # Extract 2 most recent report links from table
            report_links = new_table['Links'].values.tolist()
            found_reports = 0

            # Iterate through rows again and click the 2 extracted links
            for row in rows:
                cols = row.find_elements(By.TAG_NAME, 'td')
                for col in cols:
                    if col.text.lstrip() == 'Adherence':
                        link = col.find_element(By.TAG_NAME, 'a')
                        if link.get_attribute('href') in report_links:
                            # When href links are found click on them to start download
                            link.click()
                            found_reports += 1
                            sleep(2)

                            # Wait for .part file to disappear to verify download completed
                            while os.path.exists(os.path.join(DOWNLOAD_DIR, 'report.xls.part')):
                                sleep(1)

                            # Rename the files
                            rename_report(found_reports)

            self.crawler_log['number_of_downloads'] = found_reports
        except Exception as e:
            raise e
        finally:
            # Shutdown the Crawler
            self.end_crawl()


class ProcessDownloads:
    sheet_name = 'Sheet1'

    def __init__(self, file_path):
        # Get the file path of the file needing to be processed
        # Set Lists and Open File
        self.file_path = file_path
        self.file_list = os.listdir(DOWNLOAD_DIR)
        self.scheduled_data_set = list()
        self.actual_data_set = list()
        self.current_wb = Workbook()
        self.file_log = {
            'file_name': os.path.basename(self.file_path),
            'row_count': 0,
        }
        self.open_file()

    def open_file(self):
        # Open the current workbook in Read Only then Process File
        self.current_wb = load_workbook(self.file_path)
        self.process_file()

    def process_file(self):
        # Get current Sheet
        current_sheet = self.current_wb.get_sheet_by_name(self.sheet_name)

        # Set Default Current Variables
        current_agent = ''
        current_agent_id = 0
        current_date = ''
        current_row = 0
        current_col = 0
        current_value = ''

        # Not Used but available if needed
        description_col = 0

        # Set Default Scheduled Cols
        scheduled_from_col = 0
        scheduled_to_col = 0
        scheduled_duration_col = 0
        scheduled_activity_col = 0

        # Set Default Actual Cols and Flags
        actual_from_flag = False
        actual_from_col = 0

        actual_to_flag = False
        actual_to_col = 0

        actual_duration_flag = False
        actual_duration_col = 0

        actual_activity_flag = False
        actual_activity_col = 0

        last_from_row = 0

        # Set Stop Flags
        stop = False
        date_stop = True
        agent_stop = True

        # Iterate through each row of the report
        self.file_log['row_count'] = current_sheet.max_row
        for i, row in enumerate(current_sheet.iter_rows()):
            # Set Current Row and reset keep flags
            current_row = i + 1
            keep_scheduled_data = False
            keep_actual_data = False

            # stop is marked true when the end of the file is found
            if stop:
                break

            # reset column variables
            scheduled_from_time = ''
            scheduled_to_time = ''
            scheduled_duration = ''
            scheduled_activity = ''
            actual_from_time = ''
            actual_to_time = ''
            actual_duration = ''
            actual_activity = ''
            description = ''

            # Iterate through each column
            for j, cell in enumerate(row):
                current_col = j + 1

                if cell.value is not None:
                    # When agent is found mark Agent as found and break loop
                    if 'Agent:' in cell.value:
                        current_agent_id = extract_number(cell.value)
                        current_agent = cell.value.replace('Agent: ' + str(current_agent_id) + ' ', '')
                        agent_stop = False
                        # Nothing else can be found on row break loop and move to next row
                        break

                    # When Date is found mark Date as found and break loop
                    if 'Date: ' in cell.value and 'Generation' not in cell.value:
                        current_date = cell.value.replace('Date: ', '')
                        last_from_row = 0
                        date_stop = False
                        # Nothing else can be found on row break loop and move to next row
                        break

                    # When Scheduled Activities is found mark flag and break loop
                    if cell.value == 'Scheduled Activities':
                        date_stop = True
                        # Nothing else can be found on row break loop and move to next row
                        break

                    # When end of agent report is found break loop
                    if cell.value == 'Total for ' + str(current_agent_id) + ' ' + current_agent:
                        agent_stop = True
                        # Nothing else can be found on row break loop and move to next row
                        break

                    # When Report Parameters found loop will and then break outer loop
                    if cell.value == 'Report Parameters':
                        stop = True
                        break

                    # Define Schedule Column Numbers
                    if 'From' in cell.value and scheduled_from_col == 0:
                        scheduled_from_col = j + 1
                    if 'To' in cell.value and scheduled_to_col == 0:
                        scheduled_to_col = j + 1
                    if 'Duration' in cell.value and scheduled_duration_col == 0:
                        scheduled_duration_col = j + 1
                    if 'Activity' in cell.value and scheduled_activity_col == 0:
                        scheduled_activity_col = j + 1

                    # Define Actual Column Numbers and Mark Flags to True
                    if 'From' in cell.value and current_col > scheduled_from_col and not actual_from_flag:
                        actual_from_col = j + 1
                        actual_from_flag = True
                    if 'To' in cell.value and current_col > scheduled_to_col and not actual_to_flag:
                        actual_to_col = j + 1
                        actual_to_flag = True
                    if 'Duration' in cell.value and current_col > scheduled_duration_col and not actual_duration_flag:
                        actual_duration_col = j + 1
                        actual_duration_flag = True
                    if 'Activity' in cell.value and current_col > scheduled_activity_col and not actual_activity_flag:
                        actual_activity_col = j + 1
                        actual_activity_flag = True

                    # Log each new row that has a From value
                    if 'From' in cell.value:
                        last_from_row = i + 1

                    # Look at both stops
                    if not date_stop and not agent_stop:
                        # Check if current col has scheduled from then store data
                        if j + 1 == scheduled_from_col and current_row > last_from_row and scheduled_from_col > 0:
                            if cell.value != '--:--':
                                try:
                                    scheduled_from_time = dt.datetime.strptime(
                                        current_date + ' ' + cell.value, '%m/%d/%y %I:%M %p')
                                    # if row has scheduled data keep the data
                                    keep_scheduled_data = True
                                except:
                                    break

                        # Check if current col has scheduled to then store data
                        if j + 1 == scheduled_to_col and current_row > last_from_row and scheduled_to_col > 0:
                            if cell.value != '--:--':
                                scheduled_to_time = dt.datetime.strptime(
                                    current_date + ' ' + cell.value, '%m/%d/%y %I:%M %p')

                        # Check if current col has scheduled activity then store data
                        if j + 1 == scheduled_activity_col and current_row > last_from_row and scheduled_activity_col > 0:
                            scheduled_activity = cell.value

                        # Check if current col has actual from then store data
                        if j + 1 == actual_from_col and current_row > last_from_row and actual_from_col > 0:
                            if cell.value != '--:--':
                                try:
                                    actual_from_time = dt.datetime.strptime(
                                        current_date + ' ' + cell.value, '%m/%d/%y %I:%M %p')
                                    # if row has actual data keep the data
                                    keep_actual_data = True
                                except:
                                    break

                        # Check if current col has actual to then store data
                        if j + 1 == actual_to_col and current_row > last_from_row and actual_to_col > 0:
                            if cell.value != '--:--':
                                actual_to_time = dt.datetime.strptime(
                                    current_date + ' ' + cell.value, '%m/%d/%y %I:%M %p')

                        # Check if current col has actual activity then store data
                        if j + 1 == actual_activity_col and current_row > last_from_row and actual_activity_col > 0:
                            actual_activity = cell.value

            # If keep flag is true, store data
            if keep_scheduled_data:
                # Create row to store
                scheduled_row = [
                    current_agent_id,
                    scheduled_activity,
                    scheduled_from_time,
                    scheduled_to_time,
                    current_agent
                ]
                # Store Row
                self.scheduled_data_set.append(scheduled_row)

            if keep_actual_data:
                # Create row to store
                actual_row = [
                    current_agent_id,
                    actual_activity,
                    actual_from_time,
                    actual_to_time,
                    current_agent
                ]
                # Store row
                self.actual_data_set.append(actual_row)

        # Close the opened Workbook
        self.current_wb.close()


def convert_files():
    global files_converted

    file_list = os.listdir(DOWNLOAD_DIR)

    for file in file_list:
        if '.xlsx' not in file \
                and file.replace('.xls', '.xlsx') not in file_list:
            file_path = os.path.join(DOWNLOAD_DIR, file)
            new_file_name = file.replace('xls', 'xlsx')
            new_file_path = os.path.join(DOWNLOAD_DIR, new_file_name)
            pyexcel.save_book_as(file_name=file_path, dest_file_name=new_file_path)
            files_converted += 1


def clear_last_14_days():
    clear_actual = """
    DELETE FROM D_POST_INSTALL.T_WS_ACTIVITY_ACTUALS
    WHERE TRUNC(ACTUAL_START, 'day') >= DATEADD('day', -14, current_date)
    """
    clear_scheduled = """
    DELETE FROM D_POST_INSTALL.T_WS_ACTIVITY_SCHEDULED
    WHERE TRUNC(SCHEDULED_START, 'day') >= DATEADD('day', -14, current_date)
    """
    db.execute_query(clear_actual)
    db.execute_query(clear_scheduled)


if __name__ == '__main__':
    # Start Script

    # Define if testing is active
    testing = True
    db = SnowflakeV2(SnowflakeConnectionHandlerV2())
    db.set_user('JDLAURET')
    db.open_connection()
    crawler = PrimaryCrawler('chrome', download_directory=DOWNLOAD_DIR)
    try:
        # Define Table Names
        actual_table = 'D_POST_INSTALL.T_WS_ACTIVITY_ACTUALS'
        scheduled_table = 'D_POST_INSTALL.T_WS_ACTIVITY_SCHEDULED'
        log_file_path = os.path.join(find_main_dir(), 'log.json')

        # Define Data List to be uploaded
        upload_actual_data = []
        upload_scheduled_data = []

        # Define Today and Date String
        today = dt.datetime.today().date()
        date_string = today.strftime('%Y-%m-%d')

        try:
            log_file = open_json(log_file_path)
        except json.JSONDecodeError as e:
            log_file = {}
        log_file[date_string] = {}

        # Clear Downloads folder
        clear_downloads()

        # Run Primary Crawler

        crawler.run_crawler()
        log_file[date_string]['crawler'] = crawler.crawler_log
        save_json(log_file_path, log_file)

        # Convert Downloaded Files from .xls to .xlsx
        files_converted = 0
        convert_files()
        log_file[date_string]['converted_files'] = files_converted
        save_json(log_file_path, log_file)

        # Go through each file in the Download Directory and process all .xlsx files
        process_files = []
        for file in os.listdir(DOWNLOAD_DIR):
            if '.xlsx' in file:
                process_files.append(ProcessDownloads(os.path.join(DOWNLOAD_DIR, file)))
                # Store object data in Data lists
        for obj in process_files:
            upload_actual_data = upload_actual_data + obj.actual_data_set
            upload_scheduled_data = upload_scheduled_data + obj.scheduled_data_set
            log_file[date_string][os.path.basename(obj.file_path)] = obj.file_log
            save_json(log_file_path, log_file)

        # Clear the last 14 days of data from both tables

        try:
            clear_last_14_days()
            log_file[date_string]['cleared_data'] = True
        except:
            log_file[date_string]['cleared_data'] = False
            pass
        # Send data from data lists to tables
        table_log = {
            'actual_table_updated': False,
            'scheduled_table_updated': False
        }
        try:
            db.insert_into_table(actual_table, upload_actual_data)
            log_file[date_string]['actual_table_updated'] = True
            save_json(log_file_path, log_file)
        except:
            log_file[date_string]['actual_table_updated'] = False
            pass

        try:
            db.insert_into_table(scheduled_table, upload_scheduled_data)
            log_file[date_string]['scheduled_table_update'] = True
            save_json(log_file_path, log_file)
        except:
            log_file[date_string]['scheduled_table_update'] = False
            pass
    except Exception as e:
        crawler.end_crawl()
        raise e
    finally:
        db.close_connection()
    # End Script

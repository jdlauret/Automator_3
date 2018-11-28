import datetime as dt
import os

import pyexcel
from openpyxl import load_workbook, Workbook
from BI.data_warehouse.connector import Snowflake
from selenium import webdriver

# Directory Variables
MAIN_DIR = os.getcwd()
DOWNLOAD_DIR = os.path.join(MAIN_DIR, 'downloads')
SOURCE_FILES_DIR = os.path.join(MAIN_DIR, 'source_files')

# File path to Chrome Driver
FF_DRIVER_PATH = r'C:\\Users\\jonathan.lauret\\Google Drive\\Projects\\Chrome Driver\geckodriver.exe'

# LOGIN_URL points to the login form
LOGIN_URL = 'https://home-c20.incontact.com/inContact/Login.aspx?' \
            'ReturnUrl=%2finContact%2fManage%2fReports%2fContactDetail.aspx%3fisActiveContact%3dtrue%' \
            '26contactid%3d39558211897&isActiveContact=true&contactid=39558211897'

# REPORT_URL points to report page
CONTACT_URL = 'https://home-c20.incontact.com/inContact/Manage/Reports/' \
              'ContactDetail.aspx?isActiveContact=true&contactid={contact_id}'

# MOVE PAYLOAD TO SEPARATE FILE WHEN TESTING IS COMPLETE
payload = {
    'username': {
        'html_name': 'ctl00$BaseContent$tbxUserName',
        'value': 'mackenzie.damavandi@vivintsolar.com'
    },
    'password': {
        'html_name': 'ctl00$BaseContent$tbxPassword',
        'value': 'ba/PMO2018'
    },
}


def extract_number(string_with_number):
    """
    Find all numbers in a string and put them in an array
    :param string_with_number: A string with numbers to be extracted
    :return: First number found in string
    """
    numbers = [int(s) for s in str.split(string_with_number) if s.isdigit()]
    return numbers[0]


def rename_report(num):
    """
    Rename report.xls files to Report - num.xls
    :param num: The number that should be put in the report name
    """
    file_list = os.listdir(DOWNLOAD_DIR)
    file_name = os.path.join(DOWNLOAD_DIR, 'report - ' + str(num) + '.xls')
    for file in file_list:
        file_path = os.path.join(DOWNLOAD_DIR, file)
        if file == 'report.xls':
            os.rename(file_path, file_name)


def clear_downloads():
    """
    Delete all files in the download list
    """
    file_list = os.listdir(DOWNLOAD_DIR)
    for file in file_list:
        file_path = os.path.join(DOWNLOAD_DIR, file)
        os.remove(file_path)


class PrimaryCrawler:
    """
    Primary Crawler is a web crawler that will download 2 reports
    from Webstations and rename the files
    """
    delay = 10
    num_reports = 3

    def __init__(self):
        # Primary Items
        self.current_line = 0
        self.skip = False

        # Define Options for Firefox
        options = webdriver.FirefoxOptions()

        if not testing:
            # Set driver to headless when not in testing mode
            options.add_argument('--headless')

        fp = webdriver.FirefoxProfile()
        # Default Preference Change stops Fingerprinting on site
        fp.DEFAULT_PREFERENCES['frozen']['dom.disable_open_during_load'] = True
        # Set the Download preferences to change the download folder location and to auto download reports
        fp.set_preference('browser.download.folderList', 2)
        fp.set_preference('browser.download.manager.showWhenStarting', False)
        fp.set_preference('browser.download.dir', DOWNLOAD_DIR)
        fp.set_preference('browser.helperApps.neverAsk.saveToDisk', 'application/vnd.ms-excel')
        # Create the Driver and Delete all cookies
        self.DRIVER = webdriver.Firefox(executable_path=FF_DRIVER_PATH, options=options, firefox_profile=fp)
        self.DRIVER.delete_all_cookies()

    def login(self):
        # Use Driver to login
        # Go to Login Page
        self.DRIVER.get(LOGIN_URL)

        # Insert Username
        uname = self.DRIVER.find_element_by_name(payload['username']['html_name'])
        uname.send_keys(payload['username']['value'])

        # Insert Password
        passw = self.DRIVER.find_element_by_name(payload['password']['html_name'])
        passw.send_keys(payload['password']['value'])

        # Click the Login button
        login_button = 'ctl00$BaseContent$btnLogin'
        self.DRIVER.find_element_by_name(login_button).click()

    def run_crawler(self):
        # Login to page
        self.login()

        for contact in contact_ids:
            contact_url = CONTACT_URL.format(contact_id=contact)
            self.DRIVER.get(contact_url)
            try:
                self.DRIVER.find_element_by_id('ctl00_PopupContentPlaceHolder_tc'
                                               'QAForms_tablActiveContactOptions_tab').click()

                self.DRIVER.find_elements_by_tag_name('button')[0].click()
                self.DRIVER.switch_to_alert().accept()
            except:
                pass

        # Shutdown the Crawler
        self.end_crawl()

    def end_crawl(self):
        # Close the Driver and upload last batch
        self.DRIVER.close()


class ProcessDownloads:
    sheet_name = 'Sheet1'

    def __init__(self, file_path):
        # Get the file path of the file needing to be processed
        # Set Lists and Open File
        self.file_path = file_path
        self.file_list = os.listdir(DOWNLOAD_DIR)
        self.scheduled_data_set = list()
        self.actual_data_set = list()
        self.current_wb = Workbook()
        self.open_file()

    def open_file(self):
        # Open the current workbook in Read Only then Process File
        self.current_wb = load_workbook(self.file_path, read_only=True)
        self.process_file()

    def process_file(self):
        # Get current Sheet
        current_sheet = self.current_wb[self.sheet_name]

        # Set Default Current Variables
        current_agent = ''
        current_agent_id = 0
        current_date = ''
        current_row = 0
        current_col = 0
        current_value = ''

        # Not Used but available if needed
        description_col = 0

        # Set Default Scheduled Cols
        scheduled_from_col = 0
        scheduled_to_col = 0
        scheduled_duration_col = 0
        scheduled_activity_col = 0

        # Set Default Actual Cols and Flags
        actual_from_flag = False
        actual_from_col = 0

        actual_to_flag = False
        actual_to_col = 0

        actual_duration_flag = False
        actual_duration_col = 0

        actual_activity_flag = False
        actual_activity_col = 0

        last_from_row = 0

        # Set Stop Flags
        stop = False
        date_stop = True
        agent_stop = True

        # Iterate through each row of the report
        for i, row in enumerate(current_sheet.iter_rows()):
            # Set Current Row and reset keep flags
            current_row = i + 1
            keep_scheduled_data = False
            keep_actual_data = False

            # stop is marked true when the end of the file is found
            if stop:
                break

            # reset column variables
            scheduled_from_time = ''
            scheduled_to_time = ''
            scheduled_duration = ''
            scheduled_activity = ''
            actual_from_time = ''
            actual_to_time = ''
            actual_duration = ''
            actual_activity = ''
            description = ''

            # Iterate through each column
            for j, cell in enumerate(row):
                current_col = j + 1

                if cell.value is not None:
                    # When agent is found mark Agent as found and break loop
                    if 'Agent:' in cell.value:
                        current_agent_id = extract_number(cell.value)
                        current_agent = cell.value.replace('Agent: ' + str(current_agent_id) + ' ', '')
                        agent_stop = False
                        # Nothing else can be found on row break loop and move to next row
                        break

                    # When Date is found mark Date as found and break loop
                    if 'Date: ' in cell.value and 'Generation' not in cell.value:
                        current_date = cell.value.replace('Date: ', '')
                        last_from_row = 0
                        date_stop = False
                        # Nothing else can be found on row break loop and move to next row
                        break

                    # When Scheduled Activities is found mark flag and break loop
                    if cell.value == 'Scheduled Activities':
                        date_stop = True
                        # Nothing else can be found on row break loop and move to next row
                        break

                    # When end of agent report is found break loop
                    if cell.value == 'Total for ' + str(current_agent_id) + ' ' + current_agent:
                        agent_stop = True
                        # Nothing else can be found on row break loop and move to next row
                        break

                    # When Report Parameters found loop will and then break outer loop
                    if cell.value == 'Report Parameters':
                        stop = True
                        break

                    # Define Schedule Column Numbers
                    if 'From' in cell.value and scheduled_from_col == 0:
                        scheduled_from_col = j + 1
                    if 'To' in cell.value and scheduled_to_col == 0:
                        scheduled_to_col = j + 1
                    if 'Duration' in cell.value and scheduled_duration_col == 0:
                        scheduled_duration_col = j + 1
                    if 'Activity' in cell.value and scheduled_activity_col == 0:
                        scheduled_activity_col = j + 1

                    # Define Actual Column Numbers and Mark Flags to True
                    if 'From' in cell.value and current_col > scheduled_from_col and not actual_from_flag:
                        actual_from_col = j + 1
                        actual_from_flag = True
                    if 'To' in cell.value and current_col > scheduled_to_col and not actual_to_flag:
                        actual_to_col = j + 1
                        actual_to_flag = True
                    if 'Duration' in cell.value and current_col > scheduled_duration_col and not actual_duration_flag:
                        actual_duration_col = j + 1
                        actual_duration_flag = True
                    if 'Activity' in cell.value and current_col > scheduled_activity_col and not actual_activity_flag:
                        actual_activity_col = j + 1
                        actual_activity_flag = True

                    # Log each new row that has a From value
                    if 'From' in cell.value:
                        last_from_row = i + 1

                    # Look at both stop
                    if not date_stop and not agent_stop:
                        # Check if current col has scheduled from then store data
                        if j + 1 == scheduled_from_col and current_row > last_from_row and scheduled_from_col > 0:
                            if cell.value != '--:--':
                                try:
                                    scheduled_from_time = dt.datetime.strptime(
                                        current_date + ' ' + cell.value, '%m/%d/%y %I:%M %p')
                                    # if row has scheduled data keep the data
                                    keep_scheduled_data = True
                                except:
                                    break

                        # Check if current col has scheduled to then store data
                        if j + 1 == scheduled_to_col and current_row > last_from_row and scheduled_to_col > 0:
                            if cell.value != '--:--':
                                scheduled_to_time = dt.datetime.strptime(
                                    current_date + ' ' + cell.value, '%m/%d/%y %I:%M %p')

                        # Check if current col has scheduled activity then store data
                        if j + 1 == scheduled_activity_col and current_row > last_from_row and scheduled_activity_col > 0:
                            scheduled_activity = cell.value

                        # Check if current col has actual from then store data
                        if j + 1 == actual_from_col and current_row > last_from_row and actual_from_col > 0:
                            if cell.value != '--:--':
                                try:
                                    actual_from_time = dt.datetime.strptime(
                                        current_date + ' ' + cell.value, '%m/%d/%y %I:%M %p')
                                    # if row has actual data keep the data
                                    keep_actual_data = True
                                except:
                                    break

                        # Check if current col has actual to then store data
                        if j + 1 == actual_to_col and current_row > last_from_row and actual_to_col > 0:
                            if cell.value != '--:--':
                                actual_to_time = dt.datetime.strptime(
                                    current_date + ' ' + cell.value, '%m/%d/%y %I:%M %p')

                        # Check if current col has actual activity then store data
                        if j + 1 == actual_activity_col and current_row > last_from_row and actual_activity_col > 0:
                            actual_activity = cell.value

            # If keep flag is true, store data
            if keep_scheduled_data:
                # Create row to store
                scheduled_row = [
                    current_agent_id,
                    scheduled_activity,
                    scheduled_from_time,
                    scheduled_to_time,
                    current_agent
                ]
                # Store Row
                self.scheduled_data_set.append(scheduled_row)

            if keep_actual_data:
                # Create row to store
                actual_row = [
                    current_agent_id,
                    actual_activity,
                    actual_from_time,
                    actual_to_time,
                    current_agent
                ]
                # Store row
                self.actual_data_set.append(actual_row)

        # Close the opened Workbook
        self.current_wb.close()


def convert_files():
    file_list = os.listdir(DOWNLOAD_DIR)
    for file in file_list:
        if '.xlsx' not in file \
                and file.replace('.xls', '.xlsx') not in file_list:
            file_path = os.path.join(DOWNLOAD_DIR, file)
            new_file_name = file.replace('xls', 'xlsx')
            new_file_path = os.path.join(DOWNLOAD_DIR, new_file_name)
            pyexcel.save_book_as(file_name=file_path, dest_file_name=new_file_path)


def clear_last_14_days():
    clear_actual = """
    DELETE FROM MACK_DAMAVANDI.T_AGENT_ACTIVITY_ACTUALS
    WHERE TRUNC(ACTUAL_START) >= TRUNC(SYSDATE - 14) 
    """
    clear_scheduled = """
    DELETE FROM MACK_DAMAVANDI.T_AGENT_ACTIVITY_SCHEDULED
    WHERE TRUNC(SCHEDULED_START) >= TRUNC(SYSDATE - 14)
    """
    execute_query(clear_actual, credentials='Private')
    execute_query(clear_scheduled, credentials='Private')


if __name__ == '__main__':
    # Start Script

    # Define if testing is active
    testing = True

    import pandas as pd

    source_file = os.path.join(SOURCE_FILES_DIR, 'ActiveContacts.csv')
    contact_ids = pd.read_csv(source_file)
    contact_ids = [x[0] for x in contact_ids.values.tolist()]

    # Run Primary Crawler
    PrimaryCrawler().run_crawler()
